#Call libraries
import openpyxl
import xlrd
import unidecode
import sys
from openpyxl.workbook import Workbook 
from openpyxl import load_workbook


#Call tables
wb_main = load_workbook("/Users/Meilis/Desktop/mapping_final_after_debug/Finalized Table before 6-10 Mapping/6-10 HS task/excel/Sample_tables/Mapping.xlsx")
ws_main = wb_main.active

#Table with 10-digits code
wb_10_code = load_workbook("/Users/Meilis/Desktop/mapping_final_after_debug/Finalized Table before 6-10 Mapping/6-10 HS task/excel/Sample_tables/Exports_HS10.xlsx")
ws_10_code = wb_10_code.active



for i in range(4, ws_main.max_row+1): #final launch
    
    #Create back-up every 250 iterations
    if i % 250 == 0:
        wb_main.save('/Users/Meilis/Desktop/mapping_final_after_debug/Finalized Table before 6-10 Mapping/6-10 HS task/excel/Sample_tables/Mapping_back_up_running.xlsx')
    
    #Set HS-6 code to search
    current_hs = ws_main.cell(row = i, column=3).value.replace('.',"")
    
    #Account for NAs
    if current_hs == "NA":
        ws_main.cell(row = i, column=10).value = "NA"
        wb_main.save('/Users/Meilis/Desktop/mapping_final_after_debug/Finalized Table before 6-10 Mapping/6-10 HS task/excel/Sample_tables/Mapping.xlsx')
        print("\n", "Oops, NA detected!")
        continue
    
    print("\n", "Looking for ", current_hs, "!", "Line:", i)
    
    if i > 4:
    #Check for repeats in HS-6 codes
        if current_hs == ws_main.cell(row = (i-1), column=3).value.replace('.',""):
            #Duplicate codes
            ws_main.cell(row = i, column=10).value = ws_main.cell(row = i-1, column=10).value
            ws_main.cell(row = i, column=11).value = ws_main.cell(row = i-1, column=11).value
            ws_main.cell(row = i, column=12).value = ws_main.cell(row = i-1, column=12).value
            #Duplicate descriptions
            ws_main.cell(row = i, column=13).value = ws_main.cell(row = i-1, column=13).value
            ws_main.cell(row = i, column=14).value = ws_main.cell(row = i-1, column=14).value
            ws_main.cell(row = i, column=15).value = ws_main.cell(row = i-1, column=15).value
            print("Repeat in HS-6 code, move to the next step!")
            continue
  
    
    #Initiate a loop over Exports file
    for j in range(2, ws_10_code.max_row+1):
        
        #If 3 cells are already filled
        if ws_main.cell(row = i, column=12).value is not None:
            print("Done with", current_hs, ".", "Moving to next HS-code!")
            break
            
        #Check whether equal codes
        if current_hs == str(ws_10_code.cell(row = j, column=1).value)[0:6]:
            
            #Keep the 10-code
            hs_10_code = str(ws_10_code.cell(row = j, column=1).value)
            #Keep description
            hs_10_code_descr = str(ws_10_code.cell(row = j, column=2).value)
            
            #Fill in 3 cells and their descriptions
            #Check whether code already exists 
            if ws_main.cell(row = i, column=10).value is None:
                
                #Add a value to main table
                ws_main.cell(row = i, column=10).value = hs_10_code
                ws_main.cell(row = i, column=13).value = hs_10_code_descr
                wb_main.save('/Users/Meilis/Desktop/mapping_final_after_debug/Finalized Table before 6-10 Mapping/6-10 HS task/excel/Sample_tables/Mapping.xlsx')
                
                #Check next element is still suits current code
                if current_hs != str(ws_10_code.cell(row = j+1, column=1).value)[0:6]:
                    print("Done with", current_hs, ".", "Moving to next HS-code!")
                    break
                
                continue
                
            elif ws_main.cell(row = i, column=11).value is None:
                
                #Add a value to main table
                ws_main.cell(row = i, column=11).value = hs_10_code
                ws_main.cell(row = i, column=14).value = hs_10_code_descr
                wb_main.save('/Users/Meilis/Desktop/mapping_final_after_debug/Finalized Table before 6-10 Mapping/6-10 HS task/excel/Sample_tables/Mapping.xlsx')
                
                #Check next element is still suits current code
                if current_hs != str(ws_10_code.cell(row = j+1, column=1).value)[0:6]:
                    print("Done with", current_hs, ".", "Moving to next HS-code!")
                    break
                    
                continue
                
            elif ws_main.cell(row = i, column=12).value is None:
                
                #Add a value to main table
                ws_main.cell(row = i, column=12).value = hs_10_code
                ws_main.cell(row = i, column=15).value = hs_10_code_descr
                wb_main.save('/Users/Meilis/Desktop/mapping_final_after_debug/Finalized Table before 6-10 Mapping/6-10 HS task/excel/Sample_tables/Mapping.xlsx')
                
                #Check next element is still suits current code
                if current_hs != str(ws_10_code.cell(row = j+1, column=1).value)[0:6]:
                    print("Done with", current_hs, ".", "Moving to next HS-code!")
                    break
                
                continue
                

        else:
            continue                 





#for final adjusted
#Call tables
wb_main = load_workbook("/Users/Meilis/Desktop/Finalized mapping before sampling/To send/Mapping.xlsx")
ws_main = wb_main.active

#Table with 10-digits code
wb_10_code = load_workbook("/Users/Meilis/Desktop/Finalized mapping before sampling/To send/Exports_HS10.xlsx")
ws_10_code = wb_10_code.active

for i in range(4, ws_main.max_row+1): #final launch
    
    #Create back-up every 250 iterations
    if i % 250 == 0:
        wb_main.save('/Users/Meilis/Desktop/Finalized mapping before sampling/To send/Mapping_back.xlsx')
    
    #Set HS-6 code to search
    current_hs = ws_main.cell(row = i, column=3).value.replace('.',"")
    
    #Account for NAs
    if current_hs == "NA":
        ws_main.cell(row = i, column=10).value = "NA"
        wb_main.save('/Users/Meilis/Desktop/Finalized mapping before sampling/To send/Mapping.xlsx')
        print("\n", "Oops, NA detected!")
        continue
    
    print("\n", "Looking for ", current_hs, "!", "Line:", i)
    
    if i > 4:
    #Check for repeats in HS-6 codes
        if current_hs == ws_main.cell(row = (i-1), column=3).value.replace('.',""):
            #Duplicate codes
            ws_main.cell(row = i, column=10).value = ws_main.cell(row = i-1, column=10).value
            ws_main.cell(row = i, column=11).value = ws_main.cell(row = i-1, column=11).value
            ws_main.cell(row = i, column=12).value = ws_main.cell(row = i-1, column=12).value
            #Duplicate descriptions
            ws_main.cell(row = i, column=13).value = ws_main.cell(row = i-1, column=13).value
            ws_main.cell(row = i, column=14).value = ws_main.cell(row = i-1, column=14).value
            ws_main.cell(row = i, column=15).value = ws_main.cell(row = i-1, column=15).value
            print("Repeat in HS-6 code, move to the next step!")
            continue
  
    
    #Initiate a loop over Exports file
    for j in range(2, ws_10_code.max_row+1):
        
        #If 3 cells are already filled
        if ws_main.cell(row = i, column=12).value is not None:
            print("Done with", current_hs, ".", "Moving to next HS-code!")
            break
            
        #Check whether equal codes
        if current_hs == str(ws_10_code.cell(row = j, column=1).value)[0:6]:
            
            #Keep the 10-code
            hs_10_code = str(ws_10_code.cell(row = j, column=1).value)
            #Keep description
            hs_10_code_descr = str(ws_10_code.cell(row = j, column=2).value)
            
            #Fill in 3 cells and their descriptions
            #Check whether code already exists 
            if ws_main.cell(row = i, column=10).value is None:
                
                #Add a value to main table
                ws_main.cell(row = i, column=10).value = hs_10_code
                ws_main.cell(row = i, column=13).value = hs_10_code_descr
                wb_main.save('/Users/Meilis/Desktop/Finalized mapping before sampling/To send/Mapping.xlsx')
                
                #Check next element is still suits current code
                if current_hs != str(ws_10_code.cell(row = j+1, column=1).value)[0:6]:
                    print("Done with", current_hs, ".", "Moving to next HS-code!")
                    break
                
                continue
                
            elif ws_main.cell(row = i, column=11).value is None:
                
                #Add a value to main table
                ws_main.cell(row = i, column=11).value = hs_10_code
                ws_main.cell(row = i, column=14).value = hs_10_code_descr
                wb_main.save('/Users/Meilis/Desktop/Finalized mapping before sampling/To send/Mapping.xlsx')
                
                #Check next element is still suits current code
                if current_hs != str(ws_10_code.cell(row = j+1, column=1).value)[0:6]:
                    print("Done with", current_hs, ".", "Moving to next HS-code!")
                    break
                    
                continue
                
            elif ws_main.cell(row = i, column=12).value is None:
                
                #Add a value to main table
                ws_main.cell(row = i, column=12).value = hs_10_code
                ws_main.cell(row = i, column=15).value = hs_10_code_descr
                wb_main.save('/Users/Meilis/Desktop/Finalized mapping before sampling/To send/Mapping.xlsx')
                
                #Check next element is still suits current code
                if current_hs != str(ws_10_code.cell(row = j+1, column=1).value)[0:6]:
                    print("Done with", current_hs, ".", "Moving to next HS-code!")
                    break
                
                continue
                

        else:
            continue                 
                
