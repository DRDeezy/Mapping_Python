import requests
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from urllib import request,response
import openpyxl
import xlrd
#pip install webdriver-manager
import unidecode
from selenium.webdriver.common.keys import Keys
# login_btn=driver.find_element
import selenium.webdriver as webdriver
# browser = webdriver.Chrome(ChromeDriverManager().install())
import sys
from openpyxl.workbook import Workbook 
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException 
from selenium.common import exceptions
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import random
import time


from selenium.common.exceptions import TimeoutException


def get_results(search_term):
    
#     #Declare variables, global
    global hs_code_1
    global hs_code_2
    global hs_code_3
    
    global description_1
    global description_2
    global description_3
    
    global hs_tickers
    
    global comment
    
#     global add_descr_1
#     global add_descr_2
    global add_descr
    
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=40,20")
    
    url = "https://www.findhs.codes/"
    browser = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
    browser.get(url)
#     Waiting to locate the element
#     WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID,"txtQuery")))
    search_box = browser.find_element_by_id("txtQuery")
    search_box.send_keys(search_term)

    search_box.send_keys(Keys.ENTER)
    
    page = requests.get(browser.current_url)
    soup = BeautifulSoup(page.content, 'html.parser')
    
    if page.status_code != 200: #check if everything is fine
        comment = ("Warning when processing : " + search_term)
        return
#     else:
#         print("Processing is success!", search_term)
    
    
#     WebDriverWait(browser, 30).until(EC.presence_of_element_located((By.ID,"hidMAResult")))
    
    raw_data = soup.find(id="hidMAResult")
    raw_data_hs = raw_data["value"]
    
    #Check that there are results of a search
    check_gtip = len(browser.find_elements_by_class_name("gtip"))
    
    global check_res #check that there are results of search, not proposal to choose a category
    
    if check_gtip == 0:
        #check that there are results of search, not proposal to choose a category
            
        if "hs_codes" in raw_data_hs:
            hs_code_1 = raw_data_hs[raw_data_hs.find("hs_codes") + 11 : raw_data_hs.find("hs_codes") + 18] #extracts first HS-code
            comment = "Perfect match"
            
            if raw_data_hs[raw_data_hs.find("hs_codes") + 18] == "-":
                hs_code_2 = raw_data_hs[raw_data_hs.find("hs_codes") + 19 : raw_data_hs.find("hs_codes") + 26] #extracts second HS-code  
               
                check_res = browser.find_elements_by_class_name('vexpert_answer_header')

                if len(check_res) != 0:
                    description_1 = browser.find_elements_by_class_name("vexpert_answer_header")
#                     description_1 = browser.find_elements_by_xpath("/html/body/form/div[6]/div/div/div[1]/div[2]/div[2]/div/div[1]/div[1]/div/div[2]/div")
                    description_1 = [elem1.text for elem1 in description_1]
                    #change for class name search in case anything stops
                    
                    add_descr = browser.find_elements_by_class_name("col-lg-9")
                    add_descr = [elem_add.text for elem_add in add_descr]
#                     add_descr = add_descr.remove("")
                    while("" in add_descr) :
                        add_descr.remove("")
            
                    if len(add_descr)>2:
                        comment = "Multiple match, recheck"
#                         description_1 = description_1[0]
                        description_1 = add_descr[0].split(":")[0].split("\"")[1]
                        description_2 = add_descr[1]
                        description_3 = add_descr[2]
                    else:
                        comment = "Multiple match"
                        description_1 = description_1[0]
#                         descriptuion_1 = description_1
                        description_2 = add_descr[1]
                        description_3 = add_descr[2]
                    
                else:
                    description_1 = browser.find_elements_by_xpath("/html/body/form/div[6]/div/div/div[1]/div[2]/div/div/ul/li[1]")
                    description_1 = [elem1.text for elem1 in description_1]
                    
                    description_2 = browser.find_elements_by_xpath("/html/body/form/div[6]/div/div/div[1]/div[2]/div/div/ul/li[2]")
                    description_2 = [elem2.text for elem2 in description_2]
                    
                    description_1 = description_1[0]
                    description_2 = description_2[0]
                    description_3 = "NA"
                    comment = "Double-perfect match"    
  
            else:
                hs_code_2 = "NA"
                hs_code_3 = "NA"
                description_1 = "As item name"
                description_2 = "NA"
                description_3 = "NA"
                
        else:
            hs_code_1 = "NA"
            hs_code_2 = "NA"
            hs_code_3 = "NA"
            description_1 = "NA"
            description_2 = "NA"
            description_3 = "NA"
            comment = "No match"
        
        
    # No perfect match, several matches    
    else: 
#         description_1 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[2]/div[1]')
        
        if len(browser.find_elements_by_xpath('//*[@id="DivResults"]/div[2]/div[1]')) != 0:
    #Parse three descriptions from locations
            description_1 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[2]/div[1]')
            description_2 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[3]/div[1]')
            description_3 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[4]/div[1]')

        elif len(browser.find_elements_by_xpath('//*[@id="DivResults"]/div[2]/div[1]')) == 0:
    #Parse three descriptions from locations
            description_1 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[3]/div[1]')
            description_2 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[4]/div[1]')
            description_3 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[5]/div[1]')
   
    #Store description block
    #description_1
        description_1 = [elem1.text for elem1 in description_1]
        description_1 = description_1[0][7:-1]
        

    #description_2
        description_2 = [elem2.text for elem2 in description_2]
        description_2 = description_2[0][7:-1]
    
    #description_3
        description_3 = [elem3.text for elem3 in description_3]
        description_3 = description_3[0][7:-1]   
    
    #Collect main tickers of categories.
        hs_tickers = browser.find_elements_by_class_name("gtip_grup4")
        hs_tickers = [elem_t_1.text for elem_t_1 in hs_tickers]

    #First main HS code
        hs_code_1 = browser.find_elements_by_id("row1_" + hs_tickers[0])
        hs_code_1 = [elem_hs.text for elem_hs in hs_code_1]
        hs_code_1 = hs_code_1[0]
    #Second main HS code
        hs_code_2 = browser.find_elements_by_id("row1_" + hs_tickers[1])
        hs_code_2 = [elem_hs.text for elem_hs in hs_code_2]
        hs_code_2 = hs_code_2[0]
    #Third main HS code
        hs_code_3 = browser.find_elements_by_id("row1_" + hs_tickers[2])
        hs_code_3 = [elem_hs.text for elem_hs in hs_code_3]
        hs_code_3 = hs_code_3[0]
        comment = "Multiple match"
    
    
    browser.close()



wb = load_workbook("/Users/Meilis/Desktop/Final_Table/Mapping_main_codes_2.xlsx")
ws = wb.active

for i in range(1, ws.max_row+1):
    sleep(5)
    try:
        #Random sleep times
        time.sleep(random.randint(10,15))

        get_results(unidecode.unidecode(ws["A" + str(i)].value))

        ws.cell(row = i, column=3).value = hs_code_1
        ws.cell(row = i, column=4).value = hs_code_2
        ws.cell(row = i, column=5).value = hs_code_3

        ws.cell(row = i, column=6).value = comment

        ws.cell(row = i, column=7).value = description_1
        ws.cell(row = i, column=8).value = description_2
        ws.cell(row = i, column=9).value = description_3

        #Back-up data
        wb.save('/Users/Meilis/Desktop/Final_Table/Mapping_main_codes_3.xlsx')
    except (IndexError, KeyError, TimeoutException, exceptions.StaleElementReferenceException):
        ws.cell(row = i, column=3).value = "ERROR"
        continue



#Correct Error and Nulls code

wb = load_workbook("/Users/Meilis/Desktop/mapping_final_after_debug/Compiled_mapping_no_errors_1.xlsx")
ws = wb.active

for i in range(1723, ws.max_row+1):
    if (ws.cell(row = i, column=3).value == "ERROR") or (ws.cell(row = i, column=3).value == ""):
        sleep(10)
        try:
            #Random sleep times
            time.sleep(random.randint(10,15))

            get_results(unidecode.unidecode(ws["A" + str(i)].value))

            ws.cell(row = i, column=3).value = hs_code_1
            ws.cell(row = i, column=4).value = hs_code_2
            ws.cell(row = i, column=5).value = hs_code_3

            ws.cell(row = i, column=6).value = comment

            ws.cell(row = i, column=7).value = description_1
            ws.cell(row = i, column=8).value = description_2
            ws.cell(row = i, column=9).value = description_3

            #Back-up data
            wb.save('/Users/Meilis/Desktop/mapping_final_after_debug/Compiled_mapping_no_errors_2.xlsx')
        except (IndexError, KeyError, TimeoutException, exceptions.StaleElementReferenceException):
            ws.cell(row = i, column=3).value = "still ERROR"
            continue
    else:
        continue


#Complex mapping

def get_complex_results(search_term_complex):
    
    url = "https://www.findhs.codes/"
    browser = webdriver.Chrome(ChromeDriverManager().install())

    browser.get(url)
    search_box = browser.find_element_by_id("txtQuery")
    search_box.send_keys(search_term_complex)

    search_box.send_keys(Keys.ENTER)
    
    global hs_codes
    global description_1
    global description_2
    global description_3
    
    page = requests.get(browser.current_url)
    soup = BeautifulSoup(page.content, 'html.parser')
    
    raw_data = soup.find(id="hidMAResult")
    raw_data_hs = raw_data["value"]
    
        #Check that there are results of a search
    check_gtip = len(browser.find_elements_by_class_name("gtip"))
    
    if check_gtip == 0:
    
        if "hs_codes" in raw_data_hs:
            hs_code_1 = raw_data_hs[raw_data_hs.find("hs_codes") + 11 : raw_data_hs.find("hs_codes") + 18] #extracts first HS-code
            description_1 = browser.find_elements_by_class_name("vexpert_answer_header")
            description_1 = [elem1.text for elem1 in description_1]
            description_1 = description_1[0]
            comment = "Perfect match"
            
            if raw_data_hs[raw_data_hs.find("hs_codes") + 18] == "-":
                hs_code_2 = raw_data_hs[raw_data_hs.find("hs_codes") + 19 : raw_data_hs.find("hs_codes") + 26] #extracts second HS-code
                description_2 = "NA"
            else:
                hs_code_2 = "NA"
                hs_code_3 = "NA"
                description_2 = "NA"
                description_3 = "NA"
                
        else:
            hs_code_1 = "NA"
            hs_code_2 = "NA"
            hs_code_3 = "NA"
            description_1 = "NA"
            description_2 = "NA"
            description_3 = "NA"
            comment = "No match"
    
    
#     description_1 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[2]/div[1]')
#     description_2 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[3]/div[1]')
#     description_3 = browser.find_elements_by_xpath('//*[@id="DivResults"]/div[4]/div[1]')

#     #Store description block
#     #description_1
#     description_1 = [elem1.text for elem1 in description_1]
#     description_1 = description_1[0][:-1]
    
#     #description_2
#     description_2 = [elem2.text for elem2 in description_2]
#     description_2 = description_2[0][:-1]
    
#     #description_3
#     description_3 = [elem3.text for elem3 in description_3]
#     description_3 = description_3[0][:-1]   
    

    
#     hs_codes = browser.find_elements_by_class_name("gtip")
#     hs_codes = [elem_hs.text for elem_hs in hs_codes]
    
    print(description_1)
#     browser.close()
